import openpyxl
import pandas as pd
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog
import os
root = tk.Tk()
root.withdraw()

print("Select Synergy 'Google Photos' export file")
file = filedialog.askopenfilename()
print("File:",file)
folderName = input("Enter name of desktop folder to look in? ")


def addColumnHeader(file,col,value):
    wb = load_workbook(file) #Load Workbook
    ws = wb.active #Worksheet
    ws[col].value = value
    wb.save(file)

# Concatenates the ID field with the file location.
def populateFileLocation(file,folderName):
    wb = load_workbook(file) #Load Workbook
    ws = wb.active #Worksheet
    for row in range (2,ws.max_row+1): 
        for col in range (4,5): # column D
            char = get_column_letter(col)
            id =  ws[get_column_letter(3) + str(row)].value
            ws[char + str(row)].value = r'C:\Users\zechaaron\OneDrive - Osseo Area Schools\Desktop\{}\{}.jpg'.format(folderName,id)
    wb.save(file)

#Export File as CSV
def convertToCSV(file,newFileName):
    fileData = pd.read_excel(file,sheet_name='QRY801')
    fileData.to_csv(newFileName,index=False)
    os.remove(file) #delete XLSX file 


import subprocess
# Copy text to clip board
def copy2clip(txt):
    cmd='echo '+txt.strip()+'|clip'
    return subprocess.check_call(cmd, shell=True)

#Main Program 
if __name__ == "__main__":
    print("Starting...")
    addColumnHeader(file,'D1',"File_Location")
    populateFileLocation(file,"photos")
    convertToCSV(file,"GooglePhotos_GAM.csv")
    #Copy GAM command to OS clipboard
    copy2clip("gam csv C:\\Users\\aszadmin\\Documents\\Google Photos\\GooglePhotos_GAM.csv gam user ~Email update photo ~File_Location")
    print("finished")

